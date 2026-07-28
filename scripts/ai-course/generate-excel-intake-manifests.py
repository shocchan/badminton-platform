#!/usr/bin/env python3
"""Phase 3P-2: Excel統合候補manifestの決定的生成。

全40シートをInventory登録し、未統合シートの全行にprovenance付きintakeStatusを与える。
第一弾4シート（権利クリア）はprimary分類まで行う。教材本体へは一切書き込まない。
同一入力→同一出力（生成日時以外）。実行: python3 scripts/ai-course/generate-excel-intake-manifests.py
"""
import openpyxl, json, hashlib, re, unicodedata, sys, time, datetime, os

ROOT = os.path.join(os.path.dirname(__file__), '..', '..')
WB_PATH = os.path.join(ROOT, 'scratchpad/source-data/foundation-learning-source.xlsx')
OUT = os.path.join(ROOT, 'docs/ai-course/production/generated')
GENERATOR_VERSION = '3p2.1'
SCHEMA_VERSION = 1

FIRST_WAVE = ['オノマトペ100集（完成版）', '複合動詞一覧', '頻出表現', '最初に覚える最低限表現']
RIGHTS_UNKNOWN = {'慣用句': '慣用句110', 'ビジネスメッセージ': 'ビジネスメッセージ67', '営業': '営業用語200'}

def norm(s):
    s = unicodedata.normalize('NFKC', str(s or '')).strip()
    return re.sub(r'\s+', ' ', s)

def core(s):
    s = re.sub(r'[（(].*?[)）]', '', norm(s))
    s = ''.join(chr(ord(c) - 0x60) if 'ァ' <= c <= 'ヶ' else c for c in s)
    return re.sub(r'[、。．.,！!？?\s]', '', s)

def sha1(s):
    return hashlib.sha1(s.encode('utf-8')).hexdigest()[:16]

def main():
    t0 = time.time()
    wb_sha = hashlib.sha256(open(WB_PATH, 'rb').read()).hexdigest()[:16]
    inv = json.load(open('/tmp/excel_inventory.json'))
    sheet_meta = {s['sheet']: s for s in inv['sheets']}
    lemmas = json.load(open('/tmp/lemmas.json'))
    by_surface = {}
    for it in lemmas:
        for key in {core(it['lemma']), core(it.get('reading', ''))}:
            if key and it['id'] not in by_surface.get(key, []):
                by_surface.setdefault(key, []).append(it['id'])
    reading_of = {it['id']: core(it.get('reading', '')) for it in lemmas}

    wb = openpyxl.load_workbook(WB_PATH, data_only=True)
    assert len(wb.sheetnames) == 40, f'sheet count changed: {len(wb.sheetnames)}'

    sheets_out, candidates, structural = [], [], []

    def sheet_state(name):
        m = sheet_meta.get(name, {})
        cls = m.get('classification', 'unknown')
        used = bool(m.get('alreadyUsedBySourceRefs'))
        if m.get('rightsStatus') == 'rights_review_required':
            return 'awaiting_rights_rewrite', 'rights_unknown（CEO権利判断待ち・独自教材へ置換予定）'
        if name in FIRST_WAVE:
            return 'first_wave_classified', '権利クリア第一弾・意味分類対象'
        if used:
            return 'already_integrated', '既存140語のsourceRefで使用済み'
        if cls == 'metadata_only':
            return 'excluded_by_explicit_rule', '計画・メモ等のメタデータシート（教材行なし）'
        if cls == 'duplicate':
            return 'duplicate_source_sheet', '完成版シートと同内容の旧版'
        return 'deferred_to_phase', '後続Phase（3P-3以降）で意味分類'

    for name in wb.sheetnames:
        ws = wb[name]
        state, reason = sheet_state(name)
        rows = []
        for i, r in enumerate(ws.iter_rows(), 1):
            vals = [norm(c.value) for c in r]
            if any(vals):
                rows.append((i, vals))
        register_rows = state in ('first_wave_classified', 'deferred_to_phase',
                                  'awaiting_rights_rewrite', 'duplicate_source_sheet')
        n_struct = n_cand = 0
        if register_rows:
            for row_no, vals in rows:
                surface, extra = pick_candidate(name, vals)
                if surface is None:
                    structural.append({'sheet': name, 'excelRow': row_no,
                                       'reason': extra or 'header_or_section'})
                    n_struct += 1
                    continue
                n_cand += 1
                cid = sha1(f'{wb_sha}|{name}|{row_no}|0')
                fp = sha1(core(surface))
                matched = by_surface.get(core(surface), [])
                cand = {
                    'sourceCandidateId': cid, 'contentFingerprint': fp,
                    'provenance': {'workbookSha16': wb_sha, 'sheet': name, 'excelRow': row_no,
                                   'sourceTextOriginal': ' | '.join(v for v in vals if v)[:300]},
                    'surface': surface, 'normalizedCore': core(surface),
                    'reading': extra.get('reading', ''), 'zh': extra.get('zh', ''),
                    'example': extra.get('example', ''), 'note': extra.get('note', ''),
                    'candidateType': extra.get('type', 'unknown'),
                    'matchedExistingIds': matched,
                    'relationship': 'awaiting_review', 'intakeStatus': '', 'reason': '',
                    'requiresHumanDecision': False, 'reviewStatus': 'draft',
                }
                classify(cand, name, state, matched, reading_of)
                candidates.append(cand)
        sheets_out.append({
            'sheet': name, 'sheetState': state, 'reason': reason,
            'rightsStatus': sheet_meta.get(name, {}).get('rightsStatus', 'internally_created'),
            'nonEmptyRows': len(rows), 'structuralRows': n_struct, 'registeredCandidates': n_cand,
            'targetPhase': ('3P-2' if state == 'first_wave_classified'
                            else '3P-5' if state == 'awaiting_rights_rewrite'
                            else '3P-3以降' if state == 'deferred_to_phase' else None),
        })

    # 内容重複（cross-sheet / 同一シート内）
    seen = {}
    for c in candidates:
        seen.setdefault(c['contentFingerprint'], []).append(c)
    dup_content = 0
    for fp, group in seen.items():
        if len(group) > 1:
            group.sort(key=lambda c: (c['provenance']['sheet'] not in FIRST_WAVE,
                                      c['provenance']['sheet'], c['provenance']['excelRow']))
            for c in group[1:]:
                # rights行は行数維持・置換追跡のためdedupで状態を奪わない
                if c['intakeStatus'] not in ('duplicate_source_row', 'awaiting_rights_rewrite'):
                    c['intakeStatus'] = 'duplicate_source_row'
                    c['reason'] = f"同内容の候補が先に登録済み（{group[0]['sourceCandidateId']}）"
                    dup_content += 1

    conflicts = [c for c in candidates if c['relationship'] == 'conflict']
    awaiting = [c for c in candidates if c['intakeStatus'].startswith('awaiting')]

    write_manifests(wb_sha, sheets_out, candidates, structural, conflicts, awaiting,
                    dup_content, t0)

def pick_candidate(sheet, vals):
    """行から候補surfaceと付随fieldを取る。構造行はNoneと理由を返す。"""
    v = vals + [''] * 8
    if sheet == 'オノマトペ100集（完成版）':
        if not v[1] or 'オノマトペ' in v[1]:
            return None, 'header'
        return v[1], {'type': 'onomatopoeia', 'note': v[2], 'example': v[3]}
    if sheet == '複合動詞一覧':
        if v[0] == '・' or (re.match(r'^\d+\.?0?$', v[0]) and not v[2]):
            return None, 'section_or_header'
        if v[1] and v[2]:
            return v[1], {'type': 'grammar', 'note': v[2], 'example': v[3]}
        return None, 'note_only'
    if sheet == '頻出表現':
        if not v[0] or v[0] == '日本語' or '表現' in v[0] and '選' in v[0]:
            return None, 'header'
        return v[0], {'type': 'expression', 'reading': v[1], 'zh': v[2]}
    if sheet == '最初に覚える最低限表現':
        if not v[2] or v[2] == '日本語':
            return None, 'header_or_section'
        t = 'expression' if (len(core(v[2])) >= 6 or 'ください' in v[2] or 'ます' in v[2]) else 'word'
        return v[2], {'type': t, 'reading': v[3], 'zh': v[4], 'example': v[5]}
    # 非第一弾: 日本語を含む最初のセルをsurface扱いで登録のみ（意味分類は後続Phase）
    def is_content(x):
        if not x or x in ('False', 'True') or len(x) > 60:
            return False
        if re.fullmatch(r'[\d.．%％:：/／\-ー・, 、]+', x):
            return False
        return bool(re.search(r'[ぁ-んァ-ヶ一-龯]', x))
    first = next((x for x in vals if is_content(x)), '')
    if not first:
        return None, 'no_japanese_content'
    return first, {'type': 'unknown'}

def classify(c, sheet, state, matched, reading_of):
    if state == 'awaiting_rights_rewrite':
        c['intakeStatus'] = 'awaiting_rights_rewrite'
        c['reason'] = 'rights_unknownシート。非採用・非削除・独自教材への置換対象'
        c['requiresHumanDecision'] = True
        return
    if state == 'duplicate_source_sheet':
        c['intakeStatus'] = 'duplicate_source_row'
        c['reason'] = '完成版シートの旧版行'
        return
    if state == 'deferred_to_phase':
        c['intakeStatus'] = 'deferred_to_phase'
        c['reason'] = '第一弾対象外。3P-3以降で意味分類'
        return
    # 第一弾の意味分類
    c['intakeStatus'] = 'classified'
    if len(matched) > 1:
        c['relationship'] = 'conflict'
        c['reason'] = f'既存複数語（{",".join(matched)}）に同程度一致'
        c['requiresHumanDecision'] = True
        c['intakeStatus'] = 'awaiting_human_semantic_review'
        return
    if matched:
        r = core(c['reading'])
        # readingがかなの場合のみ比較する（最低限表現シートのreading列は拼音ローマ字のため比較不能）
        comparable = bool(re.search(r'[ぁ-ん]', r))
        if comparable and reading_of.get(matched[0]) and r != reading_of[matched[0]]:
            c['relationship'] = 'conflict'
            c['reason'] = f'表記一致だがreading不一致（候補{r}≠既存{reading_of[matched[0]]}）'
            c['requiresHumanDecision'] = True
            c['intakeStatus'] = 'awaiting_human_semantic_review'
        else:
            c['relationship'] = 'reuse_existing'
            c['reason'] = '既存語と同表記。sense統合かrelationかは人間判断'
            c['requiresHumanDecision'] = True
        return
    c['relationship'] = {'onomatopoeia': 'new_item', 'grammar': 'new_grammar_pattern',
                         'expression': 'expression', 'word': 'new_item'}.get(c['candidateType'], 'new_item')
    c['reason'] = '既存140語に一致なし'

def write_manifests(wb_sha, sheets_out, candidates, structural, conflicts, awaiting, dup_content, t0):
    from collections import Counter
    now = datetime.datetime.now().strftime('%Y-%m-%d')
    base = {'schemaVersion': SCHEMA_VERSION, 'generatedAt': now,
            'generatorVersion': GENERATOR_VERSION, 'workbookSha16': wb_sha}
    st = Counter(c['intakeStatus'] for c in candidates)
    fw = [c for c in candidates if c['provenance']['sheet'] in FIRST_WAVE]
    rel = Counter(c['relationship'] for c in fw)
    summary = {**base,
        'sheets': {'total': len(sheets_out),
                   'byState': dict(Counter(s['sheetState'] for s in sheets_out))},
        'rowCountNote': 'nonEmptyはtrim後の非空行。3A監査の4417はセル単位計上で、空白のみ行を含む',
        'rows': {'nonEmptyTotal': sum(s['nonEmptyRows'] for s in sheets_out),
                 'structural': len(structural),
                 'registeredCandidates': len(candidates)},
        'intakeStatus': dict(st), 'unclassified': sum(1 for c in candidates if not c['intakeStatus']),
        'duplicateContent': dup_content,
        'firstWave': {'sheets': FIRST_WAVE, 'candidates': len(fw), 'byRelationship': dict(rel)},
        'provenance': {'complete': sum(1 for c in candidates if c['provenance']['excelRow'] > 0
                                       and c['provenance']['sheet']), 'errors': 0},
        'generationSeconds': round(time.time() - t0, 1),
    }
    summary['provenance']['errors'] = len(candidates) - summary['provenance']['complete']
    dq = {}
    for c in awaiting:
        dq.setdefault(c['intakeStatus'], []).append(
            {'sourceCandidateId': c['sourceCandidateId'], 'surface': c['surface'],
             'sheet': c['provenance']['sheet'], 'reason': c['reason']})
    files = {
        'excel-intake-inventory.json': {**base, 'sheets': sheets_out,
                                        'structuralRows': structural},
        'content-candidates.json': {**base, 'candidates': candidates},
        'candidate-classification-summary.json': summary,
        'candidate-conflicts.json': {**base, 'conflicts': [
            {'sourceCandidateId': c['sourceCandidateId'], 'surface': c['surface'],
             'sheet': c['provenance']['sheet'], 'excelRow': c['provenance']['excelRow'],
             'matchedExistingIds': c['matchedExistingIds'], 'conflictReason': c['reason'],
             'humanDecisionRequired': True} for c in conflicts]},
        'candidate-decision-queue.json': {**base,
            'totals': {k: len(v) for k, v in dq.items()}, 'queues': dq},
    }
    for name, data in files.items():
        json.dump(data, open(os.path.join(OUT, name), 'w'), ensure_ascii=False, indent=1)
    print(json.dumps({'summary': {k: v for k, v in summary.items()
                                  if k not in ('generatedAt',)}}, ensure_ascii=False, indent=1))

if __name__ == '__main__':
    main()
